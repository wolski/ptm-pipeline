"""Shared utilities for creating small, committable test datasets from full DEA output.

Used by create_test_*.py scripts to filter phospho/protein xlsx, parquet, and
annotation files into small subsets suitable for CI testing.
"""

import csv
import random
import shutil
from collections import Counter
from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def read_xlsx_sheet(wb: openpyxl.Workbook, sheet_name: str) -> tuple[list, list[list]]:
    """Read a sheet into (headers, rows) where rows is list of lists."""
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    return list(all_rows[0]), [list(r) for r in all_rows[1:]]


def write_xlsx(sheets: dict[str, tuple[list, list[list]]], output_path: Path):
    """Write {sheet_name: (headers, rows)} to xlsx."""
    wb = openpyxl.Workbook()
    first = True
    for name, (headers, rows) in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(output_path)


def stratified_sample(items_by_group: dict[str, list], total_n: int) -> list:
    """Sample proportionally from each group, maintaining ratios."""
    grand_total = sum(len(v) for v in items_by_group.values())
    if grand_total == 0:
        return []
    selected = []
    for group, items in items_by_group.items():
        n = max(1, round(total_n * len(items) / grand_total))
        n = min(n, len(items))
        selected.extend(random.sample(items, n))
    return selected


# ---------------------------------------------------------------------------
# Site / protein selection
# ---------------------------------------------------------------------------

def select_phosphosites(
    phospho_wb: openpyxl.Workbook, n_phospho: int, keep_contrasts: list[str],
) -> tuple[set, set]:
    """Select n_phospho sites starting from diff_exp_analysis (long format).

    Strategy:
      1. Filter diff_exp_analysis to keep_contrasts
      2. Pivot: site -> {contrast: FDR} to see which sites have data in which contrasts
      3. Pick n_complete sites present in ALL contrasts (non-NA FDR)
      4. Pick n_partial sites present in only SOME contrasts (split evenly)
      5. Filter to valid sites (no REV/CON, valid SequenceWindow)

    Returns (keep_sites, keep_protein_ids).
    """
    # Read diff_exp_analysis (long: one row per site x contrast)
    headers, rows = read_xlsx_sheet(phospho_wb, "diff_exp_analysis")
    site_idx = headers.index("site")
    contrast_idx = headers.index("contrast")
    fdr_idx = headers.index("FDR")

    # Build site -> set of contrasts with non-NA FDR
    site_contrasts: dict[str, set] = {}
    for row in rows:
        contrast = row[contrast_idx]
        if contrast not in keep_contrasts:
            continue
        site = row[site_idx]
        fdr = row[fdr_idx]
        site_contrasts.setdefault(site, set())
        if fdr is not None:
            site_contrasts[site].add(contrast)

    # Read wide sheet for validation (REV, CON, SequenceWindow, modAA, protein_Id)
    w_headers, w_rows = read_xlsx_sheet(phospho_wb, "diff_exp_analysis_wide")
    w_site_idx = w_headers.index("site")
    w_rev_idx = w_headers.index("REV")
    w_con_idx = w_headers.index("CON")
    w_sw_idx = w_headers.index("SequenceWindow")
    w_modaa_idx = w_headers.index("modAA")
    w_prot_idx = w_headers.index("protein_Id")

    # Build site metadata from wide sheet
    site_meta: dict[str, dict] = {}
    for row in w_rows:
        site = row[w_site_idx]
        if row[w_rev_idx] in (True, "True", "TRUE"):
            continue
        if row[w_con_idx] in (True, "True", "TRUE"):
            continue
        sw = row[w_sw_idx]
        if not sw or str(sw) == "None" or len(str(sw)) < 7:
            continue
        site_meta[site] = {
            "modAA": str(row[w_modaa_idx]),
            "protein_Id": row[w_prot_idx],
        }

    # Split into complete (all contrasts) vs partial
    n_contrasts = len(keep_contrasts)
    complete_sites = []
    partial_sites: dict[str, list] = {c: [] for c in keep_contrasts}

    for site, contrasts in site_contrasts.items():
        if site not in site_meta:
            continue  # skip REV/CON/bad SequenceWindow
        if len(contrasts) == n_contrasts:
            complete_sites.append(site)
        else:
            # Partial: record which contrast(s) it IS present in
            for c in contrasts:
                partial_sites[c].append(site)

    print(f"Valid sites: {len(complete_sites)} complete (all {n_contrasts} contrasts), "
          f"partial: { {c: len(v) for c, v in partial_sites.items()} }")

    # Select: n_phospho complete, then add ~20% partial on top
    n_complete = min(n_phospho, len(complete_sites))
    n_partial = min(n_phospho // 5, sum(len(v) for v in partial_sites.values()))

    selected_sites = set(random.sample(complete_sites, n_complete))

    # Add partial sites, split evenly per contrast
    n_per_contrast = max(1, n_partial // n_contrasts)
    for c, sites in partial_sites.items():
        available = [s for s in sites if s not in selected_sites]
        n_pick = min(n_per_contrast, len(available))
        if n_pick > 0:
            selected_sites.update(random.sample(available, n_pick))

    keep_proteins = {site_meta[s]["protein_Id"] for s in selected_sites}

    # Report modAA distribution
    modaa_sel = Counter(site_meta[s]["modAA"] for s in selected_sites)
    n_actual_partial = len(selected_sites) - n_complete
    print(f"Selected {len(selected_sites)} phosphosites from {len(keep_proteins)} proteins")
    print(f"  {n_complete} complete + {n_actual_partial} partial")
    print(f"  modAA distribution: {dict(modaa_sel)}")

    return selected_sites, keep_proteins


def select_proteins(
    protein_wb: openpyxl.Workbook, phospho_protein_ids: set,
) -> set:
    """Select proteins: all that match the phospho protein IDs."""
    headers, rows = read_xlsx_sheet(protein_wb, "diff_exp_analysis_wide")
    prot_idx = headers.index("protein_Id")
    all_protein_set = {row[prot_idx] for row in rows}

    keep = phospho_protein_ids & all_protein_set
    print(f"Phospho proteins found in protein data: {len(keep)}")
    print(f"Phospho proteins NOT in protein data: {len(phospho_protein_ids - all_protein_set)}")

    return keep


# ---------------------------------------------------------------------------
# XLSX filtering
# ---------------------------------------------------------------------------

def _update_summary(sheets: dict) -> None:
    """Recompute the summary sheet based on filtered diff_exp_analysis_wide."""
    hdrs = sheets["diff_exp_analysis_wide"][0]
    data = sheets["diff_exp_analysis_wide"][1]
    n_total = len(data)
    rev_idx = hdrs.index("REV") if "REV" in hdrs else None
    con_idx = hdrs.index("CON") if "CON" in hdrs else None
    rev_count = sum(
        1 for r in data
        if rev_idx is not None and r[rev_idx] in (True, "True", "TRUE")
    ) if rev_idx is not None else 0
    con_count = sum(
        1 for r in data
        if con_idx is not None and r[con_idx] in (True, "True", "TRUE")
    ) if con_idx is not None else 0
    pct_con = round(100 * con_count / n_total, 2) if n_total else 0
    pct_rev = round(100 * rev_count / n_total, 2) if n_total else 0
    sheets["summary"] = (
        ["totalNrOfProteins", "percentOfContaminants", "percentOfFalsePositives",
         "NrOfProteinsNoDecoys"],
        [[n_total, pct_con, pct_rev, n_total - rev_count - con_count]],
    )


def filter_phospho_xlsx(
    wb: openpyxl.Workbook, keep_sites: set, keep_contrasts: list[str]
) -> dict[str, tuple[list, list[list]]]:
    """Filter all phospho xlsx sheets."""
    sheets = {}

    for sheet_name in wb.sheetnames:
        headers, rows = read_xlsx_sheet(wb, sheet_name)

        if sheet_name in ("annotation", "formula"):
            sheets[sheet_name] = (headers, rows)

        elif sheet_name == "summary":
            sheets[sheet_name] = (headers, rows)

        elif sheet_name == "contrasts":
            cn_idx = headers.index("contrast_name")
            filtered = [r for r in rows if r[cn_idx] in keep_contrasts]
            sheets[sheet_name] = (headers, filtered)

        elif sheet_name == "diff_exp_analysis":
            site_idx = headers.index("site")
            contrast_idx = headers.index("contrast")
            filtered = [
                r for r in rows
                if r[site_idx] in keep_sites and r[contrast_idx] in keep_contrasts
            ]
            sheets[sheet_name] = (headers, filtered)

        elif sheet_name in ("diff_exp_analysis_wide",
                            "normalized_abundances",
                            "raw_abundances_matrix",
                            "normalized_abundances_matrix",
                            "missing_information",
                            "stats_normalized",
                            "stats_normalized_wide",
                            "stats_raw",
                            "stats_raw_wide"):
            site_idx = headers.index("site")
            filtered = [r for r in rows if r[site_idx] in keep_sites]
            sheets[sheet_name] = (headers, filtered)

        else:
            sheets[sheet_name] = (headers, rows)

    _update_summary(sheets)
    return sheets


def filter_protein_xlsx(
    wb: openpyxl.Workbook, keep_proteins: set, keep_contrasts: list[str]
) -> dict[str, tuple[list, list[list]]]:
    """Filter all protein xlsx sheets."""
    sheets = {}

    for sheet_name in wb.sheetnames:
        headers, rows = read_xlsx_sheet(wb, sheet_name)

        if sheet_name in ("annotation", "formula"):
            sheets[sheet_name] = (headers, rows)

        elif sheet_name == "summary":
            sheets[sheet_name] = (headers, rows)

        elif sheet_name == "contrasts":
            cn_idx = headers.index("contrast_name")
            filtered = [r for r in rows if r[cn_idx] in keep_contrasts]
            sheets[sheet_name] = (headers, filtered)

        elif sheet_name == "diff_exp_analysis":
            prot_idx = headers.index("protein_Id")
            contrast_idx = headers.index("contrast")
            filtered = [
                r for r in rows
                if r[prot_idx] in keep_proteins and r[contrast_idx] in keep_contrasts
            ]
            sheets[sheet_name] = (headers, filtered)

        elif sheet_name in ("diff_exp_analysis_wide",
                            "normalized_abundances",
                            "raw_abundances_matrix",
                            "normalized_abundances_matrix",
                            "missing_information",
                            "stats_normalized",
                            "stats_normalized_wide",
                            "stats_raw",
                            "stats_raw_wide"):
            prot_idx = headers.index("protein_Id")
            filtered = [r for r in rows if r[prot_idx] in keep_proteins]
            sheets[sheet_name] = (headers, filtered)

        else:
            sheets[sheet_name] = (headers, rows)

    _update_summary(sheets)
    return sheets


# ---------------------------------------------------------------------------
# Parquet filtering
# ---------------------------------------------------------------------------

def filter_parquet(
    src_path: Path, dest_path: Path, filter_col: str, keep_values: set
):
    """Filter a parquet file to rows where filter_col is in keep_values."""
    table = pq.read_table(src_path)
    col = table.column(filter_col)
    mask = pa.array([v.as_py() in keep_values for v in col])
    filtered = table.filter(mask)
    pq.write_table(filtered, dest_path)
    print(f"  Parquet {src_path.name}: {len(table)} -> {len(filtered)} rows")


# ---------------------------------------------------------------------------
# Annotation TSV filtering
# ---------------------------------------------------------------------------

def filter_annotation_tsv(
    src_path: Path, dest_path: Path, keep_contrasts: list[str] | None
):
    """Filter annotation TSV, handling two formats:

    1. ContrastName/Contrast format: set unwanted contrasts to NA
    2. Group/Control format (no ContrastName): copy as-is
    """
    with open(src_path, newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = reader.fieldnames
        rows = list(reader)

    has_contrast_name = "ContrastName" in fieldnames

    with open(dest_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            if has_contrast_name and keep_contrasts is not None:
                cn = row.get("ContrastName", "").strip()
                if cn and cn != "NA" and cn not in keep_contrasts:
                    row["ContrastName"] = "NA"
                    row["Contrast"] = "NA"
            writer.writerow(row)

    fmt = "ContrastName/Contrast" if has_contrast_name else "Group/Control"
    print(f"  {src_path.name}: {len(rows)} rows ({fmt} format)")


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def run_subset(cfg: dict):
    """Run the full subsetting pipeline.

    cfg keys:
        seed, src_dir, out_dir,
        phospho_dea, protein_dea,
        phospho_res, protein_res,
        phospho_inp, protein_inp,
        phospho_xlsx, protein_xlsx,
        phospho_annot, protein_annot,  # filenames of annotation TSV in Inputs dir
        keep_contrasts, n_phospho,
    """
    random.seed(cfg["seed"])

    src_dir = Path(cfg["src_dir"])
    out_dir = Path(cfg["out_dir"])

    print(f"Source: {src_dir}")
    print(f"Output: {out_dir}")
    print()

    if not src_dir.exists():
        raise FileNotFoundError(f"Source directory not found: {src_dir}")

    phospho_xlsx_path = src_dir / cfg["phospho_dea"] / cfg["phospho_res"] / cfg["phospho_xlsx"]
    protein_xlsx_path = src_dir / cfg["protein_dea"] / cfg["protein_res"] / cfg["protein_xlsx"]

    # Step 1: Select phosphosites (from diff_exp_analysis, contrast-aware)
    print("Loading xlsx files...")
    phospho_wb = openpyxl.load_workbook(phospho_xlsx_path, read_only=True)
    print("\n--- Selecting phosphosites ---")
    keep_sites, phospho_proteins = select_phosphosites(
        phospho_wb, cfg["n_phospho"], cfg["keep_contrasts"],
    )
    phospho_wb.close()

    # Step 2: Select proteins (all matching phospho proteins)
    print("\n--- Selecting proteins ---")
    protein_wb = openpyxl.load_workbook(protein_xlsx_path, read_only=True)
    keep_proteins = select_proteins(protein_wb, phospho_proteins)
    protein_wb.close()

    # Clean and create output directories
    if out_dir.exists():
        shutil.rmtree(out_dir)

    phospho_res_out = out_dir / cfg["phospho_dea"] / cfg["phospho_res"]
    phospho_inp_out = out_dir / cfg["phospho_dea"] / cfg["phospho_inp"]
    protein_res_out = out_dir / cfg["protein_dea"] / cfg["protein_res"]
    protein_inp_out = out_dir / cfg["protein_dea"] / cfg["protein_inp"]
    for d in [phospho_res_out, phospho_inp_out, protein_res_out, protein_inp_out]:
        d.mkdir(parents=True, exist_ok=True)

    # Step 3: Filter and write phospho xlsx
    print("\n--- Filtering phospho xlsx ---")
    phospho_wb = openpyxl.load_workbook(phospho_xlsx_path, read_only=True)
    phospho_sheets = filter_phospho_xlsx(phospho_wb, keep_sites, cfg["keep_contrasts"])
    phospho_wb.close()
    write_xlsx(phospho_sheets, phospho_res_out / cfg["phospho_xlsx"])
    print(f"  Wrote {phospho_res_out / cfg['phospho_xlsx']}")
    for name, (_, rows) in phospho_sheets.items():
        print(f"    {name}: {len(rows)} rows")

    # Step 4: Filter and write protein xlsx
    print("\n--- Filtering protein xlsx ---")
    protein_wb = openpyxl.load_workbook(protein_xlsx_path, read_only=True)
    protein_sheets = filter_protein_xlsx(protein_wb, keep_proteins, cfg["keep_contrasts"])
    protein_wb.close()
    write_xlsx(protein_sheets, protein_res_out / cfg["protein_xlsx"])
    print(f"  Wrote {protein_res_out / cfg['protein_xlsx']}")
    for name, (_, rows) in protein_sheets.items():
        print(f"    {name}: {len(rows)} rows")

    # Step 5: Filter parquet files
    print("\n--- Filtering parquet files ---")
    filter_parquet(
        src_dir / cfg["phospho_dea"] / cfg["phospho_res"] / "lfqdata_normalized.parquet",
        phospho_res_out / "lfqdata_normalized.parquet",
        "site",
        keep_sites,
    )
    filter_parquet(
        src_dir / cfg["protein_dea"] / cfg["protein_res"] / "lfqdata_normalized.parquet",
        protein_res_out / "lfqdata_normalized.parquet",
        "protein_Id",
        keep_proteins,
    )

    # Step 6: Copy lfqdata.yaml files
    print("\n--- Copying lfqdata.yaml files ---")
    shutil.copy2(
        src_dir / cfg["phospho_dea"] / cfg["phospho_res"] / "lfqdata.yaml",
        phospho_res_out / "lfqdata.yaml",
    )
    shutil.copy2(
        src_dir / cfg["protein_dea"] / cfg["protein_res"] / "lfqdata.yaml",
        protein_res_out / "lfqdata.yaml",
    )
    print("  Copied lfqdata.yaml for phospho and protein")

    # Step 7: Filter annotation TSV files
    print("\n--- Filtering annotation TSV ---")
    filter_annotation_tsv(
        src_dir / cfg["phospho_dea"] / cfg["phospho_inp"] / cfg["phospho_annot"],
        phospho_inp_out / cfg["phospho_annot"],
        cfg["keep_contrasts"],
    )
    filter_annotation_tsv(
        src_dir / cfg["protein_dea"] / cfg["protein_inp"] / cfg["protein_annot"],
        protein_inp_out / cfg["protein_annot"],
        cfg["keep_contrasts"],
    )

    # Report sizes
    print("\n--- Output sizes ---")
    total_size = 0
    for f in out_dir.rglob("*"):
        if f.is_file():
            size = f.stat().st_size
            total_size += size
            print(f"  {f.relative_to(out_dir)}: {size / 1024:.1f} KB")
    print(f"\n  TOTAL: {total_size / 1024 / 1024:.1f} MB")

    print("\nDone!")
